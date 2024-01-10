#!/usr/bin/env python
#
###############################################################################
# File: LicensePlateGenerator.py
# Author: Markus Russold, markusrussold@gmail.com
# Date: 2023-05-24
# Description: This source code has been developed for and is used
#              in the experimental implementation of the paper "Incremental
#              Whole Plate ALPR Under Data Availability Constraints"
###############################################################################

import os
import random
import string
import datetime
import uuid
from collections import Counter

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from PIL import Image, ImageDraw, ImageFont, ImageFilter, ImageOps
from openpyxl import load_workbook

# Configuration parameters for execution
# Consult README for further  details
number_of_plates_to_generate = 3
number_of_images_per_plate = 2
likelyhood_to_appear = 1
show_inline = False
max_rotation_angle = 2
max_gaussianblur_radius = 2
grains_to_add = True
different_lighting = True
dest_folder = 'data'

# Function to generate a random number and returning True in case the number is lower or equal
# the likelyhood_to_appear parameter.
def random_bool():
    return random.random() <= likelyhood_to_appear

# Function to generate UUID
def generate_unique_id():
    return str(uuid.uuid4())

# Function to return specific alphabet used for the purpose of creating license plate numbers
def get_alphabet_array():
    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                'Ä', 'Ü', 'Ö']
    return alphabet

# Function to generate a random string from a given alphabet in a given length
def generate_random_string(length: int, alphabet: string):
    return ''.join(random.choices(alphabet, k=length))
    
# Function to generate a single license plate number based on a specifc syntax
def generate_plate():
    first = generate_random_string(2, get_alphabet_array())    
    second = generate_random_string(2, get_alphabet_array())    
    third = generate_random_string(4, string.digits)    
    return first + "   " + second + " " + third

# Function to remove spaces within a given string
def remove_spaces(s):
    return s.replace(" ", "")

# Function to check if an element already exists within a given list
def check_for_duplicates(mylist: [], tosearch: string):
    for item in mylist:
        if item == tosearch:
            return True
    return False

# Function inspired from
# https://stackoverflow.com/questions/53131926/i-want-to-create-salt-and-pepper-noise-function-pil-and-numpy
#  Function adds grains onto the image with a given density in order to simulate disturbances
def add_grains(image, density):
    arr = np.asarray(image)
    intensity_levels = 2 ** (arr[0, 0].nbytes * 8)

    min_intensity = 0
    max_intensity = intensity_levels - 1

    rand_image_arr = np.random.choice([min_intensity, 1, np.nan], p=[density / 2, 1 - density, density / 2], size=arr.shape)

    result_arr = arr.astype(np.float64) * rand_image_arr
    result_arr = np.nan_to_num(result_arr, nan=max_intensity).astype(arr.dtype)

    return Image.fromarray(result_arr)

# Check if the destination folder exists and create it in case it doesn't
if not os.path.exists(dest_folder):
    os.makedirs(dest_folder)

plates = []
filenames = []

# Generate the configured number of unique license plate numbers and create array
while len(plates) < number_of_plates_to_generate:
    new = generate_plate()
    if not check_for_duplicates(plates, new):
        plates.append(generate_plate())

# Define some parameters of the image
plate_width, plate_height = 520 * 2, 110 * 2
border_size = 20 * 2
frame_size = 10 * 2
background_color_plate = (255, 255, 255)
background_color_car_rand = random.randint(100,255)
background_color_car = (background_color_car_rand, background_color_car_rand, background_color_car_rand)

# Define font to be used to write the plate image
plate_font = "EuroPlate.ttf"

# Define EU patch file
eu_patch_image = Image.open('eu_patch.png')
eu_patch_image.thumbnail((100, plate_height - border_size + 1))

# Define police patch file
police_patch_image = Image.open('german_police_registration.jpg')
police_patch_image.thumbnail((100, plate_height - border_size + 1))

random.seed()
item_counter = 0

# For each produced license plate number  
for item in plates:
    item_filenames = []
    item_counter += 1
    print('License plate #', item_counter)
    
    # For all the images to be generated for a given license plate.
    # Each execution of this loop covers one column in the result matrix.
    for i in range(number_of_images_per_plate):
        if random_bool():
            font = ImageFont.truetype(plate_font, 190)
            
            text_width, text_height = font.getsize(item)
            
            # Create an image with a specific size
            image = Image.new('RGB', (plate_width + (2 * border_size), plate_height + (2 * border_size)), color = background_color_car)
            draw = ImageDraw.Draw(image)
            
            # Draw the boundaries of the license plate within the image
            draw.rounded_rectangle([(border_size, border_size), (plate_width + border_size, plate_height + border_size)], 20, fill=(0, 0, 0), outline="black")
            draw.rounded_rectangle([(border_size + frame_size, border_size + frame_size), (plate_width + border_size - frame_size, plate_height + border_size - frame_size)], 20, fill=(255, 255, 255), outline="white")
            
            # Copy the EU patch as well as the police patch into the image
            image.paste(eu_patch_image, (border_size + frame_size + 0, border_size + frame_size + 0))
            image.paste(police_patch_image, (border_size + frame_size + 290, border_size + frame_size + 0))
            
            # Put the number plate into the image
            draw.text(((plate_width / 2) - (text_width / 2) + 80, (plate_height / 2) - (text_height / 2) + 30), item, font=font, fill=(0, 0, 0))
        
            # Rotate image in order to increase disturbance
            image = image.rotate(random.randint(max_rotation_angle * -1, max_rotation_angle), expand=True, fillcolor = background_color_car)
         
            # Modify lighning conditions in order to increase disturbance
            if different_lighting:
                image = Image.fromarray(np.uint8(image))
                image = Image.eval(image, lambda x: x + 30 * random.normalvariate(0, 2))
            
            # Add blurring in order to increase disturbance
            image = image.filter(ImageFilter.GaussianBlur(radius = random.randint(0, max_gaussianblur_radius)))
    
            # Add grains in order to increase disturbance
            if grains_to_add:
                image = add_grains(image, round(random.uniform(0.0, 0.15), 2))
            
            # Convert to grayscale to simulate infrared images
            image = ImageOps.grayscale(image)
            
            # Save the generated number plate image
            filename = generate_unique_id() + ".jpg"
            img_path = dest_folder + "/" + filename
            image.save(img_path)
            item_filenames.append(filename)
            
            if show_inline:
                np_image = np.array(image)
                plt.imshow(np_image)
                plt.show()
        else:
            item_filenames.append('')
    filenames.append(item_filenames)

# Reduce any spaces to only 1 within the whole array
plates = [remove_spaces(s) for s in plates]

# Create a Pandas DataFrame with the filenames as columns and the license plate numbers as rows
df = pd.DataFrame(filenames, index=plates)

total_seen_elements = []
total_seen_elements_count = []
total_images = []

for column_name, column_data in df.iteritems():
    for index, value in column_data.iteritems():
        if value != '':
            total_seen_elements.append(index)
            total_images.append(value)
    total_seen_elements_count.append(len(set(total_seen_elements)))

# Variable counter holds a dictionary with all indexes and the number of their occurrences
counter = Counter(total_seen_elements)
counts = []

for element, count in counter.items():
    counts.append(count)

df_data_counts = df.replace('', None).count()
df_data_counts_statistics = df_data_counts.describe()
df_count = pd.DataFrame([counter]).transpose()
df_count_statistics = df_count[0].describe()
df_total_seen_elements_count = pd.DataFrame([total_seen_elements_count])
df_images = pd.DataFrame([total_images]).transpose()

# Get the current date and time and format the date and time as a string in a specific format
now = datetime.datetime.now()
filename_leading = now.strftime("0_LPN_Data_%Y-%m-%d_%H-%M")

# Save the DataFrames to an Excel file
writer = pd.ExcelWriter(dest_folder + '/' + filename_leading + '.xlsx')

df.to_excel(writer, index=True, sheet_name='data')
df_data_counts.to_excel(writer, index=True, sheet_name='data_counts')
df_data_counts_statistics.to_excel(writer, index=True, sheet_name='data_counts_stats')
df_count.to_excel(writer, index=True, sheet_name='lpn_counts')
df_count_statistics.to_excel(writer, index=True, sheet_name='lpn_counts_stats')
df_total_seen_elements_count.to_excel(writer, index=True, sheet_name='total_seen_elements_count')
df_images.to_excel(writer, index=True, sheet_name='images')
    
writer.save()

# Print a message to confirm that the file was saved
print("File saved as " + filename_leading + '.xlsx')

# Re-Open workbook (with different library) to conduct some changes and additions

# Change cell A1 to 'lpn' in order to have the index easyly accessible using that string
workbook = load_workbook(dest_folder + '/' + filename_leading + '.xlsx')
worksheet = workbook['data']
worksheet['A1'] = 'lpn'

# Create a summary worksheet and save further important information
sum_worksheet = workbook.create_sheet('summary')
sum_worksheet['A1'].value = 'Total rows (license plates):'
sum_worksheet['B1'].value = number_of_plates_to_generate
sum_worksheet['A2'].value = 'Total cols (evaluation cycles):'
sum_worksheet['B2'].value = number_of_images_per_plate
sum_worksheet['A3'].value = 'Total number of images:'
sum_worksheet['B3'].value = len(set(total_images))

sum_worksheet['A4'].value = 'Saturation at'
try:
    sum_worksheet['B4'].value = str(total_seen_elements_count.index(number_of_plates_to_generate) + 1) + ' = ' + str((total_seen_elements_count.index(number_of_plates_to_generate) + 1) / number_of_images_per_plate * 100) + '%'
except Exception:
    sum_worksheet['B4'].value = 'not saturated'
   
sum_worksheet['A6'].value = 'max_rotation_angle:'
sum_worksheet['B6'].value = max_rotation_angle
sum_worksheet['A7'].value = 'max_gaussianblur_radius:'
sum_worksheet['B7'].value = max_gaussianblur_radius
sum_worksheet['A8'].value = 'grains_to_add:'
sum_worksheet['B8'].value = grains_to_add
sum_worksheet['A9'].value = 'different_lighting:'
sum_worksheet['B9'].value = different_lighting
sum_worksheet['A10'].value = 'likelyhood_to_appear:'
sum_worksheet['B10'].value = likelyhood_to_appear

# Save the workbook
workbook.save(dest_folder + '/' + filename_leading + '.xlsx')


#######################################################
#
# The following section produces a couple of statistical
# diagrams to be further used to describe the generated
# data.
#
#######################################################

# Create a plot to visualize the saturation point
plt.plot(total_seen_elements_count)
line = plt.axvline(x=total_seen_elements_count.index(number_of_plates_to_generate) + 1, color='red', linestyle='--')
plt.savefig(dest_folder + '/' + filename_leading + '_sat_01.png', dpi=600)
x_position = line.get_xdata()[0] + 1
y_position = number_of_plates_to_generate / 2
plt.text(x_position, y_position, 'Saturation after ' + str((total_seen_elements_count.index(number_of_plates_to_generate) + 1)) + ' rounds', color=line.get_color(), ha='left', va='center', fontsize=12)
plt.savefig(dest_folder + '/' + filename_leading + '_sat_02.png', dpi=600)
if show_inline:
    plt.show()
plt.clf()

# Create various boxplots
ax = df_data_counts.plot.box()
ax.set_xticklabels(['Number of seen LPNs per cycle'])
plt.savefig(dest_folder + '/' + filename_leading + '_bp_01_01.png', dpi=600)
if show_inline:
    plt.show()
plt.clf()

ax = df_data_counts.plot.box()
ax.set_xticklabels([''])
plt.savefig(dest_folder + '/' + filename_leading + '_bp_01_02.png', dpi=600)
if show_inline:
    plt.show()
plt.clf()

ax = df_data_counts.plot.box()
ax.set_ylim(0, number_of_plates_to_generate)
ax.set_xticklabels([''])
plt.savefig(dest_folder + '/' + filename_leading + '_bp_01_03.png', dpi=600)
if show_inline:
    plt.show()
plt.clf()

ax = df_count.plot.box()
ax.set_xticklabels(['Number of occurences per LPN'])
plt.savefig(dest_folder + '/' + filename_leading + '_bp_02_01.png', dpi=600)
if show_inline:
    plt.show()
plt.clf()

ax = df_count.plot.box()
ax.set_xticklabels([''])
plt.savefig(dest_folder + '/' + filename_leading + '_bp_02_02.png', dpi=600)
if show_inline:
    plt.show()
plt.clf()

ax = df_count.plot.box()
ax.set_ylim(0, number_of_images_per_plate)
ax.set_xticklabels([''])
plt.savefig(dest_folder + '/' + filename_leading + '_bp_02_03.png', dpi=600)
if show_inline:
    plt.show()
plt.clf()